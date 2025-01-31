import cv2
from ultralytics import YOLO
from sort import Sort  # Ensure the Sort class is correctly imported
import datetime
import time
import os
from openpyxl import Workbook, load_workbook
import math
import cvzone
import numpy as np

# Get the source names
source_from = input("Name of the pond/nurseries (from): ")
source_to = input("Name of the pond/nurseries (to): ")
fish_size = input("Size of fish (in cm): ")

class FishCounter:
    def __init__(self, model_path, video_path, output_path, output_path1):
        self.model = YOLO(model_path)
        self.class_list = ['fish']
        self.tracker = Sort(max_age=1, min_hits=1, iou_threshold=0.1)  # Initialize SORT tracker with max age for quick forgetting, YOU CAN FINE TUNE THIS SECTION TO WORK OUT
        self.video_path = video_path
        self.output_path = output_path
        self.output_path1 = output_path1 #This is for to generate just footage without any detection so we can test in other program as well if you are doing live recording otherwise remove output_path1 from all
        self.counter_up = set()
        self.total_length = 0
        self.total_weight = 0
        self.object_count = 0
        self.fish_data = []
        self.line = [540, 460, 800, 460]  # Define the counting line Sometimes you might need to change it when you are working in horizontal movement or vertical. Just manually change it.
        #self.line = [550, 440, 790, 440]

    def calculate_weight(self, theta_deg, l_theta, b_theta, dpk):
        theta_rad = math.radians(theta_deg)
        cos_theta = math.cos(theta_rad)
        sin_theta = math.sin(theta_rad)
        cos_2theta = math.cos(2 * theta_rad)

        if theta_deg == 45 and l_theta == b_theta:
            l = l_theta * math.sqrt(2) / 1.29
            b = b_theta * math.sqrt(2) / 1.29 * 0.29
        else:
            b = (cos_theta * l_theta - sin_theta * b_theta) / cos_2theta
            l = (-sin_theta * l_theta + cos_theta * b_theta) / cos_2theta
        
        volume = l * b * l * dpk
        weight = 1.05 * volume
        
        return l, b, volume, weight

    def calculate_for_both_thetas(self, theta_deg, l_theta, b_theta, dpk):
        l1, b1, volume1, weight1 = self.calculate_weight(theta_deg, l_theta, b_theta, dpk)
        l2, b2, volume2, weight2 = self.calculate_weight(90 - theta_deg, b_theta, l_theta, dpk)
        
        return {
            "theta": theta_deg,
            "l_theta": (l1, l2),
            "b_theta": (b1, b2),
            "volume": (volume1, volume2),
            "weight": (weight1, weight2)
        }

    def process_frame(self, frame):
        results = self.model(frame, stream=True)
        detections = np.empty((0, 5))
    
        # Define Region of Interest (ROI) - Adjust coordinates as needed, as this is most important part if we have a plan to count fish in fixed certain area.s
        roi_x1, roi_y1, roi_x2, roi_y2 = 530, 410, 800, 610  # Example coordinates for ROI
         
        # Draw ROI on the frame for visualization
        cv2.rectangle(frame, (roi_x1, roi_y1), (roi_x2, roi_y2), (0, 255, 0), 4)

        for info in results:
            boxes = info.boxes
            for box in boxes:
                x1, y1, x2, y2 = box.xyxy[0].cpu().numpy()  # Convert tensor to numpy
                conf = box.conf[0].cpu().item()  # Convert tensor to scalar
                classindex = int(box.cls[0].cpu().item())  # Convert tensor to scalar
                objectdetect = self.class_list[classindex]
                
                # Calculate centroid of the detection
                cx, cy = (x1 + x2) // 2, (y1 + y2) // 2

                # Filter for 'fish' class and check if centroid is within ROI
                if objectdetect == 'fish' and conf > 0.2 and roi_x1 < cx < roi_x2 and roi_y1 < cy < roi_y2:
                    x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                    new_detections = np.array([x1, y1, x2, y2, conf])
                    detections = np.vstack((detections, new_detections))
                    w, h = abs(x2 - x1), abs(y2 - y1)

                    # Draw bounding box within ROI
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 100, 255), 2)
                   # cv2.rectangle(frame, (x1, y1), (x1 + w, y1 + h), (100, 255, 255), 2)

        # Update tracker with filtered detections
        track_result = self.tracker.update(detections)

        # Draw counting line
        cv2.line(frame, (self.line[0], self.line[1]), (self.line[2], self.line[3]), (0, 0, 255), 10)

        for results in track_result:
            x1, y1, x2, y2, id = results
            x1, y1, x2, y2, id = int(x1), int(y1), int(x2), int(y2), int(id)
            cx, cy = x1 + (x2 - x1) // 2, y1 + (y2 - y1) // 2

            # Draw tracked object if within ROI
            if roi_x1 < cx < roi_x2 and roi_y1 < cy < roi_y2:
                cv2.circle(frame, (cx, cy), 20, (127, 0, 0), -2)
                cvzone.putTextRect(frame, f'{id}', [x1 + 8, y1 - 12], colorR=(0, 0, 255), thickness=2, scale=1.5)

                # Check if object crossed the counting line
                if self.line[0] < cx < self.line[2] and (cy -50) < self.line[1] < (cy + 160):
                #if self.line[0] < cx < self.line[2] and (self.line[1]-50) < cy < (self.line[1] + 160):
                    cv2.line(frame, (self.line[0], self.line[1]), (self.line[2], self.line[3]), (0, 255, 0), 15)
                    
                    if id not in self.counter_up:
                        self.counter_up.add(id)
                        length_cm, weight_gm = self.calculate_fish_properties(x1, y1, x2, y2)
                        self.total_length += length_cm
                        self.total_weight += weight_gm
                        self.object_count += 1
                        self.fish_data.append((length_cm, weight_gm))

                        label = f'L: {length_cm:.2f} cm, W: {weight_gm:.2f} gm'
                        cvzone.putTextRect(frame, label, [x1, y1 - 30], colorR=(0, 255, 0), thickness=2, scale=1)
                        cvzone.putTextRect(frame, f'{objectdetect} {int(conf*100)}% L: {length_cm:.2f} cm & W: {weight_gm:.2f} gm', 
                                           [x1, y1], colorR=(0, 0, 255), thickness=2, scale=1)

        # Display the total count of fish
        cvzone.putTextRect(frame, f'Detected fish No = {len(self.counter_up)}', [80, 34], 
                           colorR=(0, 0, 255), thickness=4, scale=2.3, border=3)
        cvzone.putTextRect(frame, f'MY FISH COUNTER', [(1040), (700)], 
                           colorR=(255, 0, 100), thickness=2, scale=1.3, border=1)
        cv2.rectangle(frame, (540,420), (800,560), (255,0,255),2)
        
        #cv2.line(frame, (cy-40, cy+100), (255,0,255),4)
        
        
        return frame

    def calculate_fish_properties(self, x1, y1, x2, y2):
        dpk = 0.118 / 2.51
        p = (x2 - x1) * 0.1089
        k = (y2 - y1) * 0.1089
        l_theta = max(k, p)
        b_theta = min(k, p)
        ratio = b_theta / l_theta

        if 0 <= ratio <= 0.2858:
            results = self.calculate_for_both_thetas(90, l_theta, b_theta, dpk)
        elif 0.2858 < ratio <= 0.3641:
            results = self.calculate_for_both_thetas(85, l_theta, b_theta, dpk)
        elif 0.3641 < ratio <= 0.4399:
            results = self.calculate_for_both_thetas(80, l_theta, b_theta, dpk)
        elif 0.4399 < ratio <= 0.5143:
            results = self.calculate_for_both_thetas(75, l_theta, b_theta, dpk)
        elif 0.5143 < ratio <= 0.5885:
            results = self.calculate_for_both_thetas(70, l_theta, b_theta, dpk)
        elif 0.5885 < ratio <= 0.6636:
            results = self.calculate_for_both_thetas(65, l_theta, b_theta, dpk)
        elif 0.6636 < ratio <= 0.7409:
            results = self.calculate_for_both_thetas(60, l_theta, b_theta, dpk)
        else:
            results = self.calculate_for_both_thetas(55, l_theta, b_theta, dpk)

        return results['l_theta'][0], results['weight'][0]

    def run(self):
        cap = cv2.VideoCapture(self.video_path) # Well if you want to use webcam use (0) or (1) depending the number of camera.
        '''#If you are interested in the IP address or online streaming then use "rtsp:\\192.168.1.1\stream" it might work sometimes
        #defining url as 
        # url="rtsp://192.168.26.160/stream"
        #cap = cv2.VideoCapture(url)#self.video_path)
        # cap = cv2.VideoCapture(self.video_path)'''
        
        '''frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))''' # use these if you counting from saved footage but if you want to get maximum quality and FPS use below and change resolution and fps as you want
        frame_width = int(cap.set(cv2.CAP_PROP_FRAME_WIDTH,1280))
        frame_height = int(cap.set(cv2.CAP_PROP_FRAME_HEIGHT,720))
        
        cap.set(cv2.CAP_PROP_FPS,60)
        fps = cap.get(cv2.CAP_PROP_FPS)
        print(f"recprd_at {fps} FPS")
         

        out = cv2.VideoWriter(self.output_path, cv2.VideoWriter_fourcc(*'mp4v'), fps, (1280,720))#(frame_width, frame_height))

        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break

            frame = self.process_frame(frame)
            out.write(frame)
            cv2.imshow("Fish Tracker", frame)

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cap.release()
        out.release()
        cv2.destroyAllWindows()

        # Calculate average length and weight
        average_length = self.total_length / self.object_count if self.object_count > 0 else 0
        average_weight = self.total_weight / self.object_count if self.object_count > 0 else 0

        self.add_data_to_excel(self.object_count, source_from, source_to, fish_size, average_length, average_weight)
        self.save_fish_data_to_excel()
        print(f"Output video saved to {self.output_path}")
        print(f"Total objects detected: {self.object_count}")
        print("Count and data exported to Excel files.")

    def add_data_to_excel(self, object_count, source_from, source_to, fish_size, average_length, average_weight):
        filename = 'dpkdata.xlsx'
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_filename = f'temp_dpkdata_{timestamp}.xlsx'

        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Deepak-Detection Count"
            ws['A1'] = "Date"
            ws['B1'] = "Source From"
            ws['C1'] = "Source To"
            ws['D1'] = "Fish Size"
            ws['E1'] = "Total Fish Count"
            ws['F1'] = "Average Length of Fish (cm)"
            ws['G1'] = "Average Weight of Fish (gm)"

        next_row = ws.max_row + 1

        ws[f'A{next_row}'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'B{next_row}'] = source_from
        ws[f'C{next_row}'] = source_to
        ws[f'D{next_row}'] = fish_size
        ws[f'E{next_row}'] = object_count
        ws[f'F{next_row}'] = round(average_length, 2)
        ws[f'G{next_row}'] = round(average_weight, 2)

        try:
            wb.save(filename)
            print(f"Data added to {filename}")
        except PermissionError:
            wb.save(temp_filename)
            print(f"Error: Unable to save data to {filename}. The file is open or locked by another application.")
            print(f"Data has been saved to a temporary file: {temp_filename}")

    def save_fish_data_to_excel(self):
        filename = 'Deepak_average.xlsx'
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_filename = f'temp_deepak_average_{timestamp}.xlsx'

        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Fish Length and Weight"
            ws['A1'] = "Date"
            ws['B1'] = "Weight (gm)"
            ws['C1'] = "Length (cm)"

        next_row = ws.max_row + 1

        for length, weight in self.fish_data:
            ws[f'A{next_row}'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws[f'B{next_row}'] = round(weight, 2)
            ws[f'C{next_row}'] = round(length, 2)
            next_row += 1

        try:
            wb.save(filename)
            print(f"Fish data saved to {filename}")
        except PermissionError:
            wb.save(temp_filename)
            print(f"Error: Unable to save data to {filename}. The file is open or locked by another application.")
            print(f"Data has been saved to a temporary file: {temp_filename}")


   

if __name__ == '__main__':
    fish_counter = FishCounter(
        model_path="lastY11m500.pt", #Replace with your trained model if you want to use YOLO pre-trained model just put one "yolov8n.pt" like this.
        video_path="rds.mp4",  # Replace with the correct path to video
        output_path="My_results.mp4",  # Replace with desired output file name
        output_path1="compare_output2.mp4"  # Replace with desired output file name
    )
    fish_counter.run()
